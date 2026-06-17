from collections import defaultdict
import io
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Alignment, Border
from typing import List, Dict, Any, Tuple
from users.models import User
from rooms.models import Room
from classes.models import SchoolClass, StudentClassHistory
from openpyxl.utils import get_column_letter


def parse_seat(seat: str):
    import re

    row_letter = re.match(r"[А-Я]+", seat).group()
    col = int(re.search(r"\d+", seat).group())

    letters = "АБВГДЕЖЗИКЛМНОПРСТУФХЦЧШЩЭЮЯ"
    row = letters.index(row_letter)

    return row, col

class SeatingService:
    @staticmethod
    def get_row_letter(row_idx: int) -> str:
        letters = "АБВГДЕЖЗИКЛМНОПРСТУФХЦЧШЩЭЮЯ"
        string = ""
        base = len(letters)
        while row_idx > 0:
            row_idx, remainder = divmod(row_idx - 1, base)
            string = letters[remainder] + string
        return string

    @staticmethod
    async def prepare_data(person_ids: List[str], room_ids: List[int], teacher_ids: List[str] | None) -> Dict[str, Any]:
        students = await User.filter(person_id__in=person_ids)
        rooms = await Room.filter(id__in=room_ids)
        if teacher_ids is not None:
            teachers = await User.filter(person_id__in=teacher_ids)
        else: 
            teachers = None
        
        class_ids = list(set([s.class_id for s in students if s.class_id]))
        classes = await SchoolClass.filter(id__in=class_ids)
        class_map = {c.id: c for c in classes}
        
        history = await StudentClassHistory.filter(user__person_id__in=person_ids)
        history_map = {}
        for h in history:
            history_map.setdefault(h.user_id, set()).add(h.school_class_id)

        return {
            "students": students,
            "rooms": rooms,
            "teachers": teachers,
            "class_map": class_map,
            "history_map": history_map
        }

    @classmethod
    async def validate_seating(cls, data: Dict[str, Any]) -> Tuple[bool, str, int, int]:
        total_students = len(data["students"])
        total_capacity = sum([r.rows * r.columns for r in data["rooms"]])
        
        if total_students > total_capacity:
            return False, "Недостаточно места в аудиториях", total_students, total_capacity
             
        teachers = data.get("teachers") or []
        use_teachers = len(teachers) > 0
        if use_teachers and len(data["rooms"]) > len(data["teachers"]):
            return False, f"Недостаточно учителей. Нужно минимум {len(data['rooms'])}, дано {len(data['teachers'])}", total_students, total_capacity
            
        return True, "ok", total_students, total_capacity

    @classmethod
    def generate_seating_plan(cls, data: Dict[str, Any]) -> List[Dict[str, Any]]:
        students = data["students"]
        rooms = sorted(data["rooms"], key=lambda r: (r.corpus, r.number))
        teachers = data["teachers"]
        class_map = data["class_map"]
        history_map = data["history_map"]

        students_by_class = {}
        for s in students:
            students_by_class.setdefault(s.class_id, []).append(s)

        corpus_capacities = {}
        for r in rooms:
            corpus_capacities[r.corpus] = corpus_capacities.get(r.corpus, 0) + (r.rows * r.columns)
            
        sorted_classes = sorted(students_by_class.items(), key=lambda x: len(x[1]), reverse=True)
        class_to_corpus = {}
        
        for class_id, cls_students in sorted_classes:
            allocated = False
            for corpus in sorted(corpus_capacities.keys(), key=lambda c: corpus_capacities[c], reverse=True):
                if corpus_capacities[corpus] >= len(cls_students):
                    class_to_corpus[class_id] = corpus
                    corpus_capacities[corpus] -= len(cls_students)
                    allocated = True
                    break
            if not allocated:
                best_corpus = max(corpus_capacities, key=corpus_capacities.get)
                class_to_corpus[class_id] = best_corpus
                corpus_capacities[best_corpus] -= len(cls_students)

        rooms_by_corpus = {}
        for r in rooms:
            rooms_by_corpus.setdefault(r.corpus, []).append(r)
            
        room_assignments = {r.id: [] for r in rooms}
        
        for class_id, cls_students in students_by_class.items():
            target_corpus = class_to_corpus.get(class_id)
            corp_rooms = rooms_by_corpus.get(target_corpus, [])
            if not corp_rooms:
                corp_rooms = rooms
                
            idx = 0
            for s in cls_students:
                attempts = 0
                while attempts < len(corp_rooms):
                    r = corp_rooms[idx % len(corp_rooms)]
                    max_cap = r.rows * r.columns
                    if len(room_assignments[r.id]) < max_cap:
                        room_assignments[r.id].append(s)
                        idx += 1
                        break
                    idx += 1
                    attempts += 1

        final_seating = []
        for r in rooms:
            room_students = room_assignments[r.id]
            grid = {}
            
            for s in room_students:
                best_seat = None
                min_penalty = float('inf')
                
                for row in range(1, r.rows + 1):
                    for col in range(1, r.columns + 1):
                        if (row, col) in grid:
                            continue
                            
                        penalty = 0
                        
                        neighbors = [
                            (row - 1, col), (row + 1, col), (row, col - 1), (row, col + 1),
                            (row - 1, col - 1), (row - 1, col + 1), (row + 1, col - 1), (row + 1, col + 1)
                        ]
                        
                        for nr, nc in neighbors:
                            neighbor = grid.get((nr, nc))
                            if not neighbor:
                                continue
                                
                            is_same_desk = (nr == row) and (
                                (col % 2 == 1 and nc == col + 1) or (col % 2 == 0 and nc == col - 1)
                            )
                            
                            if is_same_desk:
                                if s.class_id == neighbor.class_id:
                                    penalty += 10**20
                                elif s.last_name == neighbor.last_name and s.middle_name[:5] == neighbor.middle_name[:5]:
                                    penalty += 10**8
                                else:
                                    penalty += 50

                            dist = max(abs(row - nr), abs(col - nc))

                            if s.class_id == neighbor.class_id:
                                if dist == 1:
                                    penalty += 10**20
                                elif dist == 2:
                                    penalty += 10**8
                                elif dist == 3:
                                    penalty += 20000
                                
                            s_class_obj = class_map.get(s.class_id)
                            n_class_obj = class_map.get(neighbor.class_id)
                            if s_class_obj and n_class_obj and s_class_obj.corpus == n_class_obj.corpus:
                                penalty += 50000
                                
                            s_hist = history_map.get(s.id, set())
                            n_hist = history_map.get(neighbor.id, set())
                            if s_hist & n_hist:
                                penalty += 100000

                        if s.sex is not None and s.class_id is not None:
                            expected_parity = (row + col) % 2
                            actual_parity = 1 if s.sex == 1 else 0
                            if expected_parity != actual_parity:
                                penalty += 1000

                        if penalty < min_penalty:
                            min_penalty = penalty
                            best_seat = (row, col)
                            
                if best_seat:
                    grid[best_seat] = s
            
            grid = cls.optimize_swap(
                grid,
                r.rows,
                r.columns,
                class_map
            )
            students_list = []
            for (row, col), s in grid.items():
                cls_obj = class_map.get(s.class_id)
                cls_name = cls_obj.display_name if cls_obj else "Неизвестен"
                fio = f"{s.last_name} {s.first_name}" + (f" {s.middle_name}" if s.middle_name else "")
                
                students_list.append({
                    "person_id": s.person_id,
                    "fio": fio,
                    "student_class": cls_name,
                    "seat": f"{SeatingService.get_row_letter(row)}{col}"
                })
                
            final_seating.append({
                "room_id": r.id,
                "corpus": r.corpus,
                "number": r.number,
                "teachers": [],
                "students": students_list,
                "_raw_students": room_students
            })

        if len(teachers) > 0:
            teacher_load = {t.id: 0 for t in teachers}

            for item in final_seating:
                room_students = item["_raw_students"]

                best_teacher = None
                best_score = float("inf")

                for t in teachers:
                    own_students_count = 0

                    for s in room_students:
                        cls_obj = class_map.get(s.class_id)
                        if cls_obj and cls_obj.teacher_id == t.id:
                            own_students_count += 1

                    score = own_students_count * 8 + teacher_load[t.id] * 20

                    if score < best_score:
                        best_score = score
                        best_teacher = t

                if best_teacher:
                    t_fio = f"{best_teacher.last_name} {best_teacher.first_name}" + (
                        f" {best_teacher.middle_name}" if best_teacher.middle_name else ""
                    )

                    item["teachers"] = [{
                        "person_id": best_teacher.person_id,
                        "fio": t_fio
                    }]

                    teacher_load[best_teacher.id] += 1
        else:
            for item in final_seating:
                item["teachers"] = []

        return final_seating

    @classmethod
    def generate_excel(cls, seating_plan: List[Dict[str, Any]]) -> io.BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Рассадка"

        font_header = Font(name="Arial", size=11, bold=True)
        font_body = Font(name="Arial", size=10)
        font_room = Font(name="Arial", size=12, bold=True)

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        center = Alignment(horizontal="center", vertical="center")

        headers = ["№", "Фамилия", "Имя", "Отчество", "Класс", "Место"]
        current_row = 1

        for room in seating_plan:
            teachers = room.get("teachers", [])
            teachers_str = ", ".join(t["fio"] for t in teachers)

            title = f"Корпус {room['corpus']} - Аудитория {room['number']}"
            if teachers_str:
                title += f" ({teachers_str})"

            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
            cell = ws.cell(row=current_row, column=1, value=title)
            cell.font = font_room
            cell.alignment = center
            ws.row_dimensions[current_row].height = 25
            current_row += 1

            for col, h in enumerate(headers, 1):
                c = ws.cell(row=current_row, column=col, value=h)
                c.font = font_header
                c.alignment = center

            current_row += 1

            sorted_students = sorted(room["students"], key=lambda x: parse_seat(x["seat"]))

            for i, s in enumerate(sorted_students, 1):
                parts = s["fio"].split()
                last_name = parts[0] if len(parts) > 0 else ""
                first_name = parts[1] if len(parts) > 1 else ""
                middle_name = parts[2] if len(parts) > 2 else ""

                row_data = [
                    i,
                    last_name,
                    first_name,
                    middle_name,
                    s["student_class"],
                    s["seat"]
                ]

                for col, val in enumerate(row_data, 1):
                    cell = ws.cell(row=current_row, column=col, value=val)
                    cell.font = font_body
                    cell.alignment = center
                    cell.border = thin_border

                ws.row_dimensions[current_row].height = 18
                current_row += 1

            current_row += 1

        for col in ws.columns:
            col_letter = openpyxl.utils.get_column_letter(col[0].column)

            if col_letter == "A":
                ws.column_dimensions[col_letter].width = 5
                continue

            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col_letter].width = max(max_len + 2, 10)

        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        return file_stream
    

    @classmethod
    def sorted_generate_excel(cls, seating_plan: List[Dict[str, Any]]) -> io.BytesIO:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Рассадка"

        font_header = Font(name="Arial", size=11, bold=True)
        font_body = Font(name="Arial", size=10)

        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        center = Alignment(horizontal="center", vertical="center")

        headers = ["№", "Корпус", "Аудитория", "Фамилия", "Имя", "Отчество", "Класс", "Место"]

        ws.append(headers)

        header_row = 1

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=col)
            cell.font = font_header
            cell.alignment = center

        row_idx = 2

        for room in seating_plan:
            corpus = room["corpus"]
            number = room["number"]

            sorted_students = sorted(room["students"], key=lambda x: parse_seat(x["seat"]))

            for i, s in enumerate(sorted_students, 1):
                parts = s["fio"].split()

                last_name = parts[0] if len(parts) > 0 else ""
                first_name = parts[1] if len(parts) > 1 else ""
                middle_name = parts[2] if len(parts) > 2 else ""

                ws.append([
                    i,
                    corpus,
                    number,
                    last_name,
                    first_name,
                    middle_name,
                    s["student_class"],
                    s["seat"]
                ])

                for col in range(1, 9):
                    cell = ws.cell(row=row_idx, column=col)
                    cell.font = font_body
                    cell.alignment = center
                    cell.border = thin_border

                row_idx += 1
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = f"A1:H{row_idx - 1}"
        for col in range(1, 9):
            col_letter = get_column_letter(col)

            if col_letter == "A":
                ws.column_dimensions[col_letter].width = 5
            else:
                max_len = 0
                for cell in ws[col_letter]:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))

                ws.column_dimensions[col_letter].width = max(max_len + 2, 10)

        file_stream = io.BytesIO()
        wb.save(file_stream)
        file_stream.seek(0)
        return file_stream
    

    @staticmethod
    def calc_class_radius(seating_plan, max_radius=5):
        results = {r: 0 for r in range(1, max_radius + 1)}

        for room in seating_plan:
            grid = {}

            for s in room["students"]:
                row, col = parse_seat(s["seat"])
                grid[(row, col)] = s

            items = list(grid.items())

            for i in range(len(items)):
                (r1, c1), s1 = items[i]

                for j in range(i + 1, len(items)):
                    (r2, c2), s2 = items[j]

                    if s1["student_class"] != s2["student_class"]:
                        continue

                    dist_r = max(abs(r1 - r2), abs(c1 - c2))

                    for radius in range(1, max_radius + 1):
                        if dist_r <= radius:
                            results[radius] += 1
                            break
        return results
    

    @staticmethod
    def optimize_swap(grid, rows, cols, class_map, iterations=2000):
        import random

        def score(g):
            total = 0

            items = list(g.items())

            for i in range(len(items)):
                (r1, c1), s1 = items[i]

                for j in range(i + 1, len(items)):
                    (r2, c2), s2 = items[j]

                    if s1.class_id != s2.class_id:
                        continue

                    dist = max(abs(r1 - r2), abs(c1 - c2))

                    if dist == 1:
                        total += 10**9
                    elif dist == 2:
                        total += 80000
                    elif dist == 3:
                        total += 20000

            return total

        best = dict(grid)
        best_score = score(best)

        for _ in range(iterations):
            if len(best) < 2:
                break

            (p1, s1), (p2, s2) = random.sample(list(best.items()), 2)

            new_grid = dict(best)
            new_grid[p1], new_grid[p2] = new_grid[p2], new_grid[p1]

            new_score = score(new_grid)

            if new_score < best_score:
                best = new_grid
                best_score = new_score

        return best
    
    @staticmethod
    def check_seating(seating_plan):
        metrics = {
            "classes_split": 0,
            "max_one_class_romm": 0
        }
        class_corpuses = defaultdict(set)

        for room in seating_plan:
            corpus = room["corpus"]

            for s in room["students"]:
                class_corpuses[s["student_class"]].add(corpus)

        metrics["classes_split"] = sum(
            1 for c, corpuses in class_corpuses.items() if len(corpuses) > 1
        )

        for room in seating_plan:
            counter = defaultdict(int)

            for s in room["students"]:
                counter[s["student_class"]] += 1

            if counter:
                metrics["max_one_class_romm"] = max(
                    metrics["max_one_class_romm"],
                    max(counter.values())
                )

        return metrics