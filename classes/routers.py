from collections import defaultdict

from fastapi import APIRouter, HTTPException, status
from fastapi import UploadFile, File
from fastapi.responses import StreamingResponse
import io
from openpyxl import Workbook, load_workbook
from tortoise.expressions import Q

from classes.models import SchoolClass, TeacherAssignment
from fastapi import Query
from subjects.models import Subject
from users.routers import transliterate
from tortoise.transactions import in_transaction
from classes.service import (
    add_student_class,
    assign_teacher_to_class,
    move_student_group,
    remove_teacher_assignment,
    remove_teacher_from_class,
    transfer_student,
    remove_student_class,
    class_students,
    create_class,
    update_class_obj,
    delete_class,
    get_class_obj,
    class_student_objects,
    class_history_ids
)
from classes.permissions import (
    can_manage_class,
    can_view_class,
    is_homeroom_teacher,
    is_operator_or_above,
)
from classes.schemas import (
    ClassCreate,
    ClassImportRow,
    ClassUpdate,
    ClassResponse,
    ClassStudentResponse,
    AssignTeacherRequest,
    FullClassResponse,
    StudentCard,
    StudentInviteRequest,
    StudentInviteResponse,
    AddExistingStudentRequest,
    MoveGroupRequest,
    TeacherAssignmentResponse,
    TeacherCard,
    TransferStudentRequest,
    AssignTeacherExtended
)
from core.config import settings
from core.permissions import min_perms
from core.roles import ROLE_TEACHER, ROLE_STUDENT
from users.models import User

router = APIRouter()


async def class_response(c: SchoolClass, students_count: int = 0) -> ClassResponse:
    students = await class_student_objects(c)

    students_g1 = []
    students_g2 = []

    for u in students:
        card = StudentCard(
            id=u.id,
            person_id=u.person_id,
            first_name=u.first_name,
            last_name=u.last_name,
            middle_name=u.middle_name,
            group=u.class_group,
        )

        if u.class_group == 1:
            students_g1.append(card)
        elif u.class_group == 2:
            students_g2.append(card)

    history = await class_history_ids(c)

    assignments = await TeacherAssignment.filter(
        school_class=c
    ).select_related("teacher")

    teachers_g1 = {}
    teachers_g2 = {}

    for a in assignments:
        t = a.teacher
        subject = a.subject

        key = (t.id, subject)

        if key not in teachers_g1:
            teachers_g1[key] = TeacherCard(
                id=t.id,
                person_id=t.person_id,
                first_name=t.first_name,
                last_name=t.last_name,
                middle_name=t.middle_name,
                groups=[],
                subject=subject,
            )

        if key not in teachers_g2:
            teachers_g2[key] = TeacherCard(
                id=t.id,
                person_id=t.person_id,
                first_name=t.first_name,
                last_name=t.last_name,
                middle_name=t.middle_name,
                groups=[],
                subject=subject,
            )

        if a.group == 1:
            teachers_g1[key].groups.append(1)
        elif a.group == 2:
            teachers_g2[key].groups.append(2)
        else:
            teachers_g1[key].groups.extend([1, 2])
            teachers_g2[key].groups.extend([1, 2])

    return ClassResponse(
        id=c.id,
        teacher_id=c.teacher_id,
        parallel=c.parallel,
        litera=c.litera,
        corpus=c.corpus,

        students_group_first=students_g1,
        students_group_second=students_g2,

        teachers_group_first=list(teachers_g1.values()),
        teachers_group_second=list(teachers_g2.values()),

        history=history,
        display_name=f"{c.parallel}{c.litera}",
        students_count=students_count,
    )


async def get_full_class(school_class: SchoolClass):
    students = await class_student_objects(school_class)

    students_group_1 = []
    students_group_2 = []

    for u in students:
        student_data = {
            "id": u.id,
            "person_id": u.person_id,
            "first_name": u.first_name,
            "last_name": u.last_name,
            "middle_name": u.middle_name,
            "email": u.email,
            "login": u.login,
            "group": u.class_group,
        }

        if u.class_group == 1:
            students_group_1.append(student_data)
        elif u.class_group == 2:
            students_group_2.append(student_data)

    assignments = await TeacherAssignment.filter(
        school_class=school_class
    ).select_related("teacher")

    homeroom_teacher = None
    subject_teachers = {}

    if school_class.teacher_id:
        t = await User.get_or_none(id=school_class.teacher_id)
        if t:
            homeroom_teacher = {
                "id": t.id,
                "person_id": t.person_id,
                "first_name": t.first_name,
                "last_name": t.last_name,
                "middle_name": t.middle_name,
            }

    teacher_map = defaultdict(lambda: {
        "id": None,
        "person_id": None,
        "first_name": None,
        "last_name": None,
        "middle_name": None,
        "groups": set()
    })

    for a in assignments:
        t = a.teacher
        key = (t.id, a.subject)

        if teacher_map[key]["id"] is None:
            teacher_map[key].update({
                "id": t.id,
                "person_id": t.person_id,
                "first_name": t.first_name,
                "last_name": t.last_name,
                "middle_name": t.middle_name,
                "subject": a.subject,
            })

        if a.group in (1, 2):
            teacher_map[key]["groups"].add(a.group)
        else:
            teacher_map[key]["groups"].update([1, 2])

    subject_teachers = [
        {**v, "groups": list(v["groups"])}
        for v in teacher_map.values()
    ]

    return {
        "id": school_class.id,
        "teacher": homeroom_teacher,
        "parallel": school_class.parallel,
        "litera": school_class.litera,
        "corpus": school_class.corpus,
        "display_name": f"{school_class.parallel}{school_class.litera}",

        "students_group_first": students_group_1,
        "students_group_second": students_group_2,

        "teachers": subject_teachers,
        "count_teachers": len(subject_teachers),

        "history": await class_history_ids(school_class)
    }


@router.get("/health")
async def health():
    return {"service": "classes", "status": "ok"}


@router.get("/show-short", response_model=list[ClassResponse])
@min_perms(settings.TEACHER_ROLE)
async def list_classes(current_user: User):
    if is_operator_or_above(current_user):
        classes = await SchoolClass.all().order_by("corpus", "parallel", "litera")
    else:
        classes = await SchoolClass.filter(
            teacher_id=current_user.id
        ).order_by("corpus", "parallel", "litera")

    result = []
    for c in classes:
        students = await class_students(c)
        result.append(await class_response(c, len(students)))

    return result

@router.get("/show-full", response_model=list[FullClassResponse])
@min_perms(settings.TEACHER_ROLE)
async def list_classes(current_user: User):
    if is_operator_or_above(current_user):
        classes = await SchoolClass.all().order_by("corpus", "parallel", "litera")
    else:
        classes = await SchoolClass.filter(
            teacher_id=current_user.id
        ).order_by("corpus", "parallel", "litera")

    result = []
    for c in classes:
        result.append(await get_full_class(c))

    return result


@router.post("/", response_model=ClassResponse, status_code=status.HTTP_201_CREATED)
@min_perms(settings.TEACHER_ROLE)
async def create_class_endpoint(current_user: User, body: ClassCreate):
    school_class = await create_class(
        parallel=body.parallel,
        litera=body.litera,
        corpus=body.corpus,
    )
    return await class_response(school_class)


@router.get("/{class_id}/short", response_model=ClassResponse)
@min_perms(settings.TEACHER_ROLE)
async def get_class(current_user: User, class_id: int):
    school_class = await get_class_obj(class_id)
    await can_view_class(current_user, school_class)

    students = await class_students(school_class)
    return await class_response(school_class, len(students))

@router.get("/{class_id}/full", response_model=FullClassResponse)
@min_perms(settings.TEACHER_ROLE)
async def get_class(current_user: User, class_id: int):
    school_class = await get_class_obj(class_id)
    await can_view_class(current_user, school_class)

    return await get_full_class(school_class)


@router.patch("/{class_id}", response_model=ClassResponse)
@min_perms(settings.TEACHER_ROLE)
async def update_class(current_user: User, body: ClassUpdate, class_id: int):
    school_class = await get_class_obj(class_id)
    can_manage_class(current_user, school_class)

    school_class = await update_class_obj(
        school_class,
        body.model_dump(exclude_unset=True),
    )

    students = await class_students(school_class)
    return await class_response(school_class, len(students))


@router.delete("/{class_id}", status_code=status.HTTP_204_NO_CONTENT)
@min_perms(settings.OPERATOR_ROLE)
async def delete_class_endpoint(
    current_user: User, 
    class_id: int, 
    delete_students: bool = Query(False, description="Удалить ли всех учащихся из этого класса?")
):
    school_class = await get_class_obj(class_id)
    
    async with in_transaction():
        if delete_students:
            await User.filter(class_id=class_id).delete()
            
        await delete_class(school_class)
        
    return None


@router.patch("/{class_id}/teacher", response_model=ClassResponse)
@min_perms(settings.OPERATOR_ROLE)
async def set_class_teacher(current_user: User, class_id: int, body: AssignTeacherRequest):
    school_class = await get_class_obj(class_id)

    teacher = await User.get_or_none(id=body.teacher_id)
    if not teacher:
        raise HTTPException(404, "Учитель не найден")

    if teacher.role != ROLE_TEACHER:
        raise HTTPException(400, "Пользователь не учитель")

    school_class.teacher_id = teacher.id
    await school_class.save()

    students = await class_students(school_class)
    return await class_response(school_class, len(students))

@router.post("/{class_id}/assign-teacher", response_model=TeacherAssignmentResponse)
@min_perms(settings.TEACHER_ROLE)
async def assign_teacher(
    class_id: int,
    body: AssignTeacherExtended,
    current_user: User,
):
    school_class = await get_class_obj(class_id)

    already_assigned = await TeacherAssignment.filter(
        teacher_id=body.teacher_id,
        school_class=school_class,
        group=body.group,
        subject=body.subject
    ).exists()

    if already_assigned:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Этот учитель уже назначен на данный предмет в этой группе/классе."
        )
    
    exists = await Subject.filter(
        id=body.subject
    ).filter(
        Q(admins__id=body.teacher_id) | Q(teachers__id=body.teacher_id)
    ).exists()

    if exists:
        assignment = await assign_teacher_to_class(
            actor=current_user,
            teacher_id=body.teacher_id,
            school_class=school_class,
            group=body.group,
            subject=body.subject
        )
    else:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Этот учитель не является учителем данного предмета"
        )

    return TeacherAssignmentResponse(
        id=assignment.id,
        teacher_id=assignment.teacher_id,
        class_id=school_class.id,
        group=assignment.group,
        subject=assignment.subject
    )

@router.get("/my/assigned-classes")
@min_perms(settings.TEACHER_ROLE)
async def my_assigned_classes(current_user: User):
    assignments = await TeacherAssignment.filter(
        teacher_id=current_user.id
    ).select_related("school_class")

    result = {}

    for a in assignments:
        c = a.school_class

        class_id = c.id

        if class_id not in result:
            result[class_id] = {
                "class_id": class_id,
                "parallel": c.parallel,
                "litera": c.litera,
                "corpus": c.corpus,
                "subjects": defaultdict(list),
            }

        result[class_id]["subjects"][a.subject].append({
            "group": a.group,
        })

    for class_data in result.values():
        class_data["subjects"] = dict(class_data["subjects"])

    return list(result.values())

@router.delete("/teacher-assignment/{assignment_id}")
@min_perms(settings.TEACHER_ROLE)
async def delete_assignment(
    assignment_id: int,
    current_user: User,
):
    await remove_teacher_assignment(
        actor=current_user,
        assignment_id=assignment_id,
    )
    return {"status": "ok"}

@router.delete("/{class_id}/teacher/{teacher_id}") # Удалит учителя вообще со всего класса!
@min_perms(settings.TEACHER_ROLE)
async def remove_teacher_from_class_endpoint(
    class_id: int,
    teacher_id: int,
    current_user: User,
):
    await remove_teacher_from_class(
        actor=current_user,
        teacher_id=teacher_id,
        class_id=class_id,
    )

    return {"status": "ok"}

@router.delete("/{class_id}/teacher", response_model=ClassResponse)
@min_perms(settings.OPERATOR_ROLE)
async def remove_class_teacher(current_user: User, class_id: int):
    school_class = await get_class_obj(class_id)

    school_class.teacher_id = None
    await school_class.save()

    students = await class_students(school_class)
    return await class_response(school_class, len(students))


@router.get("/{class_id}/students", response_model=list[ClassStudentResponse])
@min_perms(settings.TEACHER_ROLE)
async def get_class_students(current_user: User, class_id: int):
    school_class = await get_class_obj(class_id)
    await can_view_class(current_user, school_class)

    return await class_students(school_class)


@router.post("/{class_id}/students", response_model=StudentInviteResponse)
@min_perms(settings.TEACHER_ROLE)
async def add_student(current_user: User, class_id: int, body: AddExistingStudentRequest):
    school_class = await get_class_obj(class_id)
    student = await User.get_or_none(id=body.user_id)

    if not student:
        raise HTTPException(404, "Пользователь не найден")

    await add_student_class(
        actor=current_user,
        school_class=school_class,
        student=student,
        group=body.group,
    )

    return StudentInviteResponse(
        user_id=student.id,
        login=student.login,
        email=student.email,
        class_id=school_class.id,
        group=body.group,
        password_link_sent=False,
    )


@router.patch("/{class_id}/students/{user_id}/group", response_model=StudentInviteResponse)
@min_perms(settings.TEACHER_ROLE)
async def move_group(current_user: User, class_id: int, user_id: int, body: MoveGroupRequest):
    school_class = await get_class_obj(class_id)
    student = await User.get_or_none(id=user_id)

    if not student:
        raise HTTPException(404, "Пользователь не найден")

    await move_student_group(
        actor=current_user,
        school_class=school_class,
        student=student,
        group=body.group,
    )

    return StudentInviteResponse(
        user_id=student.id,
        login=student.login,
        email=student.email,
        class_id=school_class.id,
        group=body.group,
        password_link_sent=False,
    )


@router.post("/{class_id}/students/{user_id}/transfer", response_model=StudentInviteResponse)
@min_perms(settings.TEACHER_ROLE)
async def transfer(
    current_user: User, 
    class_id: int,
    user_id: int,
    body: TransferStudentRequest
):
    source_class = await get_class_obj(class_id)
    target_class = await get_class_obj(body.target_class_id)
    student = await User.get_or_none(id=user_id)

    if not student:
        raise HTTPException(404, "Пользователь не найден")

    await transfer_student(
        actor=current_user,
        source_class=source_class,
        target_class=target_class,
        student=student,
        group=body.group,
    )

    return StudentInviteResponse(
        user_id=student.id,
        login=student.login,
        email=student.email,
        class_id=target_class.id,
        group=body.group,
        password_link_sent=False,
    )


@router.delete("/{class_id}/students/{user_id}", status_code=status.HTTP_204_NO_CONTENT)
@min_perms(settings.TEACHER_ROLE)
async def remove_student(current_user: User, class_id: int, user_id: int):
    school_class = await get_class_obj(class_id)
    student = await User.get_or_none(id=user_id)

    if not student:
        raise HTTPException(404, "Пользователь не найден")

    await remove_student_class(
        actor=current_user,
        school_class=school_class,
        student=student,
    )

    return None


@router.get("/{class_id}/export")
@min_perms(settings.TEACHER_ROLE)
async def export_class(class_id: int, current_user: User = None, group: int | None = None):
    school_class = await get_class_obj(class_id)
    await can_view_class(current_user, school_class)

    students = await class_students(school_class)

    if group:
        students = [s for s in students if s["group"] == group]

    wb = Workbook()
    ws = wb.active
    ws.title = "students"

    ws.append(["email", "first_name", "last_name", "group"])

    for s in students:
        ws.append([
            s["email"],
            s["first_name"],
            s["last_name"],
            s["group"],
        ])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=class_{class_id}.xlsx"
        },
    )


@router.post("/{class_id}/import")
@min_perms(settings.OPERATOR_ROLE)
async def import_class(
    class_id: int,
    current_user: User = None, 
    file: UploadFile = File(...),
    dry_run: bool = False,
):
    school_class = await get_class_obj(class_id)
    can_manage_class(current_user, school_class)

    if not file.filename.endswith(".xlsx"):
        raise HTTPException(400, "Only .xlsx allowed")

    wb = load_workbook(io.BytesIO(await file.read()))
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    if len(rows) < 2:
        raise HTTPException(400, "Empty file")

    header = [h.lower() for h in rows[0]]

    required = {"email", "first_name", "last_name", "group"}
    if not required.issubset(set(header)):
        raise HTTPException(400, f"Missing columns: {required}")

    idx = {k: i for i, k in enumerate(header)}

    parsed = []
    errors = []

    for i, row in enumerate(rows[1:], start=2):
        try:
            item = ClassImportRow(
                email=row[idx["email"]],
                first_name=row[idx["first_name"]],
                last_name=row[idx["last_name"]],
                group=row[idx["group"]],
            )
            parsed.append(item)
        except Exception as e:
            errors.append({"row": i, "error": str(e)})

    if errors:
        return {
            "status": "failed",
            "errors": errors,
        }

    emails = [r.email for r in parsed]
    if len(emails) != len(set(emails)):
        raise HTTPException(400, "Duplicate emails in file")

    if dry_run:
        return {
            "status": "ok",
            "rows": len(parsed),
        }

    created = 0

    async with in_transaction():
        for r in parsed:
            user = await User.get_or_none(email=r.email)

            if not user:
                last_latin = transliterate(r.last_name).capitalize()
                first_initial = transliterate(r.first_name[0]).upper() if r.first_name else ""
                
                base_login = f"{last_latin}{first_initial}"
                generated_login = base_login
                
                counter = 2
                while await User.exists(login=generated_login):
                    generated_login = f"{base_login}{counter}"
                    counter += 1

                user = await User.create(
                    email=r.email,
                    login=generated_login,
                    first_name=r.first_name,
                    last_name=r.last_name,
                    role=1,
                    class_id=school_class.id,
                    class_group=r.group,
                    password_hash="UNSET_PASSWORD_CHOSEN_BY_EMAIL",
                    must_set_password=True,
                )
                
                from auth.password_service import build_password_link, mark_user_must_set_password
                from mail.gmail_service import send_password_email

                raw_token = await mark_user_must_set_password(user)
                link = build_password_link(raw_token)
                try:
                    await send_password_email(
                        to=user.email,
                        subject="Активация аккаунта",
                        greeting=f"Здравствуйте, {user.last_name} {user.first_name}!",
                        link=link,
                    )
                except Exception:
                    pass

            await add_student_class(
                actor=current_user,
                school_class=school_class,
                student=user,
                group=r.group,
            )

            created += 1

    return {
        "status": "ok",
        "imported": created,
    }